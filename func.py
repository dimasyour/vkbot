import datetime
from openpyxl import load_workbook
from datetime import datetime as dt

import openpyxl
wb = openpyxl.load_workbook(filename='raspisanie.xlsx')
# Функция, чтобы узнать какая учебная неделя из экселя 
def checkWeek():
    wb1 = load_workbook('weeks.xlsx')
    sheet_ranges = wb['sheet1']
    column_d = sheet_ranges['D']
    column_c = sheet_ranges['C']
    weeksTodayStr = datetime.datetime.now()
    weekTodayStr = weeksTodayStr.strftime("%U")
    weekTodayInt = int(weekTodayStr)
    weekTodayInt = weekTodayInt+0
    for i in range(len(column_c)):
        if  column_d[i].value == weekTodayInt:
            week = column_c[i].value
            break
        elif column_d[i].value == None:
            break
    return str(week)
# Функция сообщения для кнопки НЕДЕЛЯ
def weekFull():
    wb1 = load_workbook('weeks.xlsx')
    sheet_ranges = wb['sheet1']
    column_d = sheet_ranges['D']
    column_c = sheet_ranges['C']
    column_g = sheet_ranges['G']
    column_h = sheet_ranges['H']
    weeksTodayStr = datetime.datetime.now()
    weekTodayStr = weeksTodayStr.strftime("%U")
    weekTodayInt = int(weekTodayStr)
    weekTodayInt = weekTodayInt+0
    for i in range(len(column_c)):
        if  column_d[i].value == weekTodayInt:
            week = column_c[i].value
            break
        elif column_d[i].value == None:
            break
    for i in range(len(column_c)):
        if  column_d[i].value == weekTodayInt:
            week1 = column_g[i].value
            break
        elif column_d[i].value == None:
            break
    for i in range(len(column_c)):
        if  column_d[i].value == weekTodayInt:
            week2 = column_h[i].value
            break
        elif column_d[i].value == None:
            break
    return 'Сечас идёт '+str(week)+'-ая неделя'+'\nНачалась: '+str(week1)+'\nЗакончится: '+str(week2)



#2 - показать сколько осталось, 1 - название пары
def para(choose):
    x = choose
    h, m = dt.now().hour, dt.now().minute
    if (((h == 8 or h == 9) and m <=30)) and (x == 1):
        return 'Первая пара'
    elif (((h == 8 or h == 9) and m <=30)) and (x == 2):   
        result = (9*60+30)-(h*60+m)
        return result
    elif ((h == 9 and m >= 40) or (h == 10) or (h == 11 and m <= 10)) and (x == 1):
        return 'Вторая пара'
    elif ((h == 9 and m >= 40) or (h == 10) or (h == 11 and m <= 10)) and (x == 2):
        result = (11*60+10)-(h*60+m)
        return result
    elif ((h == 11 and m >= 40) or (h == 12) or (h == 13 and m <= 10)) and (x == 1):
        return 'Третья пара'
    elif ((h == 11 and m >= 40) or (h == 12) or (h == 13 and m <= 10)) and (x == 2):
        result = (13*60+10)-(h*60+m)
        return result
    elif ((h == 13 and m >= 30) or (h == 14) or (h == 15 and m <= 0)) and (x == 1):
        return 'Четвертая пара'
    elif ((h == 13 and m >= 30) or (h == 14) or (h == 15 and m <= 0)) and (x == 2):
        result = (15*60+0)-(h*60+m)
        return result
    elif ((h == 15 and m >= 10) or (h == 16 and m <= 40)) and (x == 1):
        return 'Пятая пара'
    elif ((h == 15 and m >= 10) or (h == 16 and m <= 40)) and (x == 2):
        result = (16*60+40)-(h*60+m)
        return result
    elif ((h == 16 and m >= 50) or (h == 17) or (h == 18 and m <= 20)) and (x == 1):
        return 'Шестая пара'
    elif ((h == 16 and m >= 50) or (h == 17) or (h == 18 and m <= 20)) and (x == 2):
        result = (18*60+20)-(h*60+m)
        return result
    elif ((h == 18 and m >= 20) or (h == 19) or (h == 20 and m <= 0)) and (x == 1):
        return 'Седьмая пара'
    elif ((h == 18 and m >= 20) or (h == 19) or (h == 20 and m <= 0)) and (x == 2):
        result = (20*60+0)-(h*60+m)
        return result

def get_mon(group):
    sheet = group
    ws = wb[sheet]
    column_mon = ws['A']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))

#get_mon('фмфи-б18пио')

def get_tue(group):
    sheet = group
    ws = wb[sheet]
    column_mon = ws['B']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))


def get_wed(group):
    sheet = group
    ws = wb[sheet]
    column_mon = ws['C']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))

def get_thu(group):
    sheet = group
    ws = wb[sheet]
    column_mon = ws['D']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))

def get_fri(group):
    sheet = group
    ws = wb[sheet]
    column_mon = ws['E']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))

def get_sab(group):
    sheet = group
    ws = wb[sheet]
    column_mon = ws['F']
    vals = []
    for i in range(len(column_mon)):
        if (column_mon[i].value != '-' and column_mon[i].value != None):
            vals.append(column_mon[i].value)
    return '\n'.join(map(str, vals))